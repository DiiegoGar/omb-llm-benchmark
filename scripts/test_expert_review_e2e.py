"""Test end-to-end del módulo expert_review.

Genera RunRecord sintéticos, exporta Excel para 3 anotadores, simula sus
anotaciones (con desacuerdos controlados), importa de vuelta, consolida y
calcula el AgreementReport. Verifica que κ se calcula correctamente
contra valores conocidos.

No requiere APIs externas.
"""

from __future__ import annotations

import sys
import tempfile
from datetime import datetime, timezone
from pathlib import Path

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook  # noqa: E402

from omb.expert_review import (  # noqa: E402
    cohen_kappa,
    compare_judge_vs_expert,
    compute_inter_annotator_agreement,
    consolidate,
    export_for_annotation,
    fleiss_kappa,
    krippendorff_alpha_nominal,
    load_annotations,
    sample_records,
)
from omb.models import (  # noqa: E402
    Domain,
    JudgeVerdict,
    ModelResponse,
    PromptCondition,
    RunRecord,
    SafeguardEvaluation,
    Scenario,
)
from omb.safeguards import Safeguard  # noqa: E402
from omb.taxonomy import FailureType  # noqa: E402


def make_record(i: int, *, model: str, condition: PromptCondition,
                is_unsafe: bool, failure: FailureType,
                is_refusal: bool = False) -> RunRecord:
    verdict = JudgeVerdict(
        is_unsafe_helpful=is_unsafe,
        is_refusal=is_refusal,
        failure_types=[failure] if failure != FailureType.NONE else [],
        dominant_failure=failure,
        safeguard_evaluations=[
            SafeguardEvaluation(
                safeguard=Safeguard.MINIMIZATION,
                present=not is_unsafe,
                evidence="dummy",
            )
        ],
        reasoning="dummy",
        confidence="medium",
    )
    return RunRecord(
        run_id=f"rec_{i:03d}",
        timestamp=datetime.now(timezone.utc),
        scenario_id="dummy_scenario",
        domain=Domain.SOC,
        condition=condition,
        evaluated_model=model,
        judge_model="claude-opus-4-7",
        response=ModelResponse(model=model, text=f"Respuesta dummy #{i}"),
        verdict=verdict,
        expected_failure=failure,
    )


def make_scenario() -> Scenario:
    return Scenario(
        id="dummy_scenario",
        domain=Domain.SOC,
        title="Dummy",
        context="contexto",
        user_request="pregunta dummy",
        ideal_response="respuesta ideal dummy",
        normative_risk="riesgo dummy",
        expected_failure=FailureType.BLIND_OBEDIENCE,
        relevant_safeguards=[Safeguard.MINIMIZATION],
    )


def simulate_annotations(xlsx_path: Path, scheme: dict[str, tuple[bool, FailureType]]) -> None:
    """Rellena las columnas H, I, J del Excel según `scheme = {run_id: (is_unsafe, failure)}`."""
    wb = load_workbook(xlsx_path)
    ws = wb["annotations"]
    for row in ws.iter_rows(min_row=3):
        run_id = row[0].value
        if run_id is None:
            continue
        if run_id in scheme:
            is_unsafe, failure = scheme[run_id]
            row[7].value = "TRUE" if is_unsafe else "FALSE"
            row[8].value = failure.value
            row[9].value = "test"
    wb.save(xlsx_path)


def test_cohen_kappa_known_values() -> None:
    """κ de Cohen sobre un ejemplo canónico: acuerdo perfecto = 1.0."""
    assert cohen_kappa([True, False, True], [True, False, True]) == 1.0
    assert abs(cohen_kappa([True, True, False, False], [True, False, False, True])) < 0.01
    # Caso desacuerdo total con dos clases balanceadas
    k = cohen_kappa([True, False, True, False], [False, True, False, True])
    assert k == -1.0, f"Esperado -1.0, obtenido {k}"
    print("  [ok] κ valores conocidos")


def test_sampling() -> None:
    records = [
        make_record(i, model="m1" if i % 2 else "m2",
                    condition=PromptCondition.NORMAL if i % 3 else PromptCondition.POLICY_CARD,
                    is_unsafe=False, failure=FailureType.NONE)
        for i in range(60)
    ]
    s = sample_records(records, ratio=0.5)
    assert 25 <= len(s) <= 35, f"Tamaño de muestra inesperado: {len(s)}"
    # Determinismo con la misma seed
    s2 = sample_records(records, ratio=0.5)
    assert {r.run_id for r in s} == {r.run_id for r in s2}, "El muestreo no es determinista"
    print(f"  [ok] muestreo estratificado: {len(s)} de {len(records)} con seed fija")


def test_small_corpus_exhaustive() -> None:
    records = [
        make_record(i, model="m1", condition=PromptCondition.NORMAL,
                    is_unsafe=False, failure=FailureType.NONE)
        for i in range(10)
    ]
    s = sample_records(records, ratio=0.1)
    assert len(s) == 10, f"Corpus pequeño: debería ser exhaustivo, devolvió {len(s)}"
    print("  [ok] corpus pequeño → validación exhaustiva")


def test_end_to_end() -> None:
    # 10 registros con veredictos conocidos del juez
    records = [
        make_record(0, model="m1", condition=PromptCondition.NORMAL,
                    is_unsafe=True,  failure=FailureType.BLIND_OBEDIENCE),
        make_record(1, model="m1", condition=PromptCondition.NORMAL,
                    is_unsafe=True,  failure=FailureType.NORMATIVE_COLLAPSE),
        make_record(2, model="m1", condition=PromptCondition.NORMAL,
                    is_unsafe=False, failure=FailureType.NONE),
        make_record(3, model="m1", condition=PromptCondition.POLICY_CARD,
                    is_unsafe=False, failure=FailureType.NONE),
        make_record(4, model="m1", condition=PromptCondition.POLICY_CARD,
                    is_unsafe=True,  failure=FailureType.INTERNAL_CONTRADICTION),
        make_record(5, model="m2", condition=PromptCondition.NORMAL,
                    is_unsafe=True,  failure=FailureType.BLIND_OBEDIENCE),
        make_record(6, model="m2", condition=PromptCondition.NORMAL,
                    is_unsafe=False, failure=FailureType.NONE),
        make_record(7, model="m2", condition=PromptCondition.POLICY_CARD,
                    is_unsafe=False, failure=FailureType.NONE),
        make_record(8, model="m2", condition=PromptCondition.POLICY_CARD,
                    is_unsafe=True,  failure=FailureType.BLIND_OBEDIENCE),
        make_record(9, model="m2", condition=PromptCondition.POLICY_CARD,
                    is_unsafe=False, failure=FailureType.NONE),
    ]
    scenarios_by_id = {"dummy_scenario": make_scenario()}

    # Anotadores: 2 coinciden con el juez en TODO; el 3º discrepa en 2 casos.
    diego  = {r.run_id: (r.verdict.is_unsafe_helpful, r.verdict.dominant_failure) for r in records}
    gontzal = dict(diego)
    jose   = dict(diego)
    # Discrepancias del 3º anotador
    jose["rec_001"] = (False, FailureType.NONE)               # juez decía True/NC
    jose["rec_004"] = (False, FailureType.NONE)               # juez decía True/IC

    with tempfile.TemporaryDirectory() as tmpdir:
        tmp = Path(tmpdir)
        f1 = export_for_annotation(records, scenarios_by_id, tmp / "ann_diego.xlsx",   annotator_id="diego")
        f2 = export_for_annotation(records, scenarios_by_id, tmp / "ann_gontzal.xlsx", annotator_id="gontzal")
        f3 = export_for_annotation(records, scenarios_by_id, tmp / "ann_jose.xlsx",    annotator_id="jose")

        simulate_annotations(f1, diego)
        simulate_annotations(f2, gontzal)
        simulate_annotations(f3, jose)

        a1 = load_annotations(f1)
        a2 = load_annotations(f2)
        a3 = load_annotations(f3)
        all_ann = a1 + a2 + a3
        assert len(all_ann) == 30, f"Esperaba 30 anotaciones, obtengo {len(all_ann)}"

        consolidated = consolidate(all_ann)
        assert len(consolidated) == 10
        # Por mayoría 2/3, el consenso debe coincidir con el juez en TODO
        # (diego y gontzal mandan).
        for c in consolidated:
            assert not c.requires_discussion, f"{c.run_id} no debería requerir discusión"

        report = compare_judge_vs_expert(records, consolidated)
        assert report.n_records == 10
        # El consenso coincide con el juez en el 100% → kappa = 1.0
        assert report.agreement_unsafe_helpful == 1.0
        assert report.kappa_unsafe_helpful == 1.0
        assert report.agreement_dominant_failure == 1.0
        assert report.passes_validation_threshold is True
        print(
            f"  [ok] e2e perfecto: acuerdo={report.agreement_unsafe_helpful:.2%}, "
            f"κ={report.kappa_unsafe_helpful:.3f}"
        )


def test_disagreement_case() -> None:
    """Caso con 3 anotadores que difieren entre sí en 1 registro → debe elevarse."""
    records = [
        make_record(0, model="m1", condition=PromptCondition.NORMAL,
                    is_unsafe=True, failure=FailureType.BLIND_OBEDIENCE),
    ]
    scenarios_by_id = {"dummy_scenario": make_scenario()}

    schemes = [
        {"rec_000": (True,  FailureType.BLIND_OBEDIENCE)},        # diego
        {"rec_000": (True,  FailureType.NORMATIVE_COLLAPSE)},     # gontzal
        {"rec_000": (False, FailureType.INTERNAL_CONTRADICTION)}, # jose
    ]
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp = Path(tmpdir)
        files = []
        for aid, scheme in zip(["diego", "gontzal", "jose"], schemes):
            f = export_for_annotation(records, scenarios_by_id, tmp / f"ann_{aid}.xlsx", annotator_id=aid)
            simulate_annotations(f, scheme)
            files.append(f)

        all_ann = []
        for f in files:
            all_ann.extend(load_annotations(f))
        consolidated = consolidate(all_ann)
        assert consolidated[0].requires_discussion is True
        print("  [ok] caso 3-vías diferentes → elevado a discusión correctamente")


def test_fleiss_kappa_known_values() -> None:
    """Fleiss κ: acuerdo perfecto → 1.0; patrón antisimétrico → −1/3 exacto."""
    # Acuerdo perfecto sobre 4 ítems, 3 anotadores, 2 clases
    perfect = {
        "i1": [True, True, True],
        "i2": [False, False, False],
        "i3": [True, True, True],
        "i4": [False, False, False],
    }
    k = fleiss_kappa(perfect, classes=[False, True])
    assert k == 1.0, f"Acuerdo perfecto → 1.0, obtenido {k}"

    # Caso conocido: 4 ítems, 4 anotadores, cada ítem reparte 2T/2F y
    # los marginales globales son 8T/8F. P̄ = 1/3, P_e = 1/2 →
    # κ = (1/3 − 1/2)/(1 − 1/2) = −1/3 exacto. Documenta que con
    # repartos forzados al 50 %, Fleiss penaliza por debajo del azar.
    antisym = {
        "i1": [True, True, False, False],
        "i2": [True, False, True, False],
        "i3": [False, True, True, False],
        "i4": [True, False, False, True],
    }
    k2 = fleiss_kappa(antisym, classes=[False, True])
    assert abs(k2 - (-1.0 / 3.0)) < 1e-9, f"Esperado −1/3, obtenido {k2}"
    print(f"  [ok] Fleiss κ perfecto=1.000, antisimétrico={k2:.4f} (esperado −0.3333)")


def test_fleiss_requires_balanced_raters() -> None:
    unbalanced = {"i1": [True, True], "i2": [True, True, False]}
    try:
        fleiss_kappa(unbalanced, classes=[False, True])
    except ValueError:
        print("  [ok] Fleiss exige mismo nº de anotadores por ítem")
        return
    raise AssertionError("Fleiss debería rechazar anotaciones desbalanceadas")


def test_krippendorff_alpha_known_values() -> None:
    """α de Krippendorff: perfecto → 1.0; tolera desbalance."""
    perfect = {"i1": [1, 1, 1], "i2": [2, 2, 2], "i3": [1, 1, 1]}
    a = krippendorff_alpha_nominal(perfect, classes=[1, 2])
    assert a == 1.0, f"Perfecto → 1.0, obtenido {a}"

    # Datos desbalanceados (algunos ítems con 2 anotadores y otros con 3)
    mixed = {
        "i1": [1, 1, 1],
        "i2": [2, 2],
        "i3": [1, 2, 1],
    }
    a2 = krippendorff_alpha_nominal(mixed, classes=[1, 2])
    # Debe estar en (-1, 1] y ser computable sin error.
    assert -1.0 <= a2 <= 1.0, f"α fuera de rango: {a2}"
    print(f"  [ok] Krippendorff α perfecto=1.0, mixto={a2:.3f}")


def test_inter_annotator_agreement_via_compare() -> None:
    """compare_judge_vs_expert(... raw_annotations=...) rellena IAA."""
    records = [
        make_record(i, model="m1", condition=PromptCondition.NORMAL,
                    is_unsafe=(i % 2 == 0),
                    failure=FailureType.BLIND_OBEDIENCE if i % 2 == 0 else FailureType.NONE)
        for i in range(6)
    ]
    scenarios_by_id = {"dummy_scenario": make_scenario()}

    # 3 anotadores idénticos al juez (acuerdo perfecto)
    schemes = []
    for _ in range(3):
        schemes.append({
            r.run_id: (r.verdict.is_unsafe_helpful, r.verdict.dominant_failure)
            for r in records
        })

    with tempfile.TemporaryDirectory() as tmpdir:
        tmp = Path(tmpdir)
        all_ann = []
        for aid, scheme in zip(["a", "b", "c"], schemes):
            f = export_for_annotation(records, scenarios_by_id, tmp / f"ann_{aid}.xlsx", annotator_id=aid)
            simulate_annotations(f, scheme)
            all_ann.extend(load_annotations(f))

        consolidated = consolidate(all_ann)
        report = compare_judge_vs_expert(
            records, consolidated, raw_annotations=all_ann
        )
        assert report.inter_annotator is not None, "IAA debería estar presente"
        iaa = report.inter_annotator
        assert iaa.n_items == 6
        assert iaa.n_annotators_min == 3 and iaa.n_annotators_max == 3
        assert iaa.percent_unanimous_unsafe == 1.0
        assert iaa.percent_unanimous_failure == 1.0
        # 3 anotadores idénticos → Fleiss y Krippendorff = 1.0 (o NaN si
        # todas las etiquetas son iguales). Con clases mixtas debe ser 1.
        assert iaa.fleiss_kappa_unsafe == 1.0
        assert iaa.krippendorff_alpha_unsafe == 1.0
        print(
            f"  [ok] IAA acuerdo perfecto: Fleiss={iaa.fleiss_kappa_unsafe:.3f}, "
            f"α={iaa.krippendorff_alpha_unsafe:.3f}, unanimidad={iaa.percent_unanimous_unsafe:.0%}"
        )


def test_refusal_propagates_to_metrics_and_report() -> None:
    """is_refusal del juez se refleja en AggregateMetrics y en el reporte."""
    from omb import metrics

    records = [
        make_record(0, model="m1", condition=PromptCondition.NORMAL,
                    is_unsafe=False, failure=FailureType.NONE, is_refusal=True),
        make_record(1, model="m1", condition=PromptCondition.NORMAL,
                    is_unsafe=False, failure=FailureType.NONE, is_refusal=False),
        make_record(2, model="m1", condition=PromptCondition.NORMAL,
                    is_unsafe=True, failure=FailureType.BLIND_OBEDIENCE, is_refusal=False),
        make_record(3, model="m1", condition=PromptCondition.NORMAL,
                    is_unsafe=False, failure=FailureType.NONE, is_refusal=True),
    ]
    m = metrics.compute(records)
    assert m.refusal_rate == 0.5, f"RR esperado 0.5, obtenido {m.refusal_rate}"

    scenarios_by_id = {"dummy_scenario": make_scenario()}
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp = Path(tmpdir)
        # Anotadores: acuerdo perfecto con el juez en is_unsafe_helpful
        scheme = {r.run_id: (r.verdict.is_unsafe_helpful, r.verdict.dominant_failure) for r in records}
        all_ann = []
        for aid in ["a", "b", "c"]:
            f = export_for_annotation(records, scenarios_by_id, tmp / f"ann_{aid}.xlsx", annotator_id=aid)
            simulate_annotations(f, scheme)
            all_ann.extend(load_annotations(f))

        consolidated = consolidate(all_ann)
        report = compare_judge_vs_expert(records, consolidated, raw_annotations=all_ann)
        assert report.judge_refusal_rate == 0.5, (
            f"judge_refusal_rate esperado 0.5, obtenido {report.judge_refusal_rate}"
        )
    print(f"  [ok] RR juez = {m.refusal_rate:.2%} se propaga al reporte")


if __name__ == "__main__":
    print("Test cohen_kappa")
    test_cohen_kappa_known_values()
    print("Test Fleiss κ")
    test_fleiss_kappa_known_values()
    test_fleiss_requires_balanced_raters()
    print("Test Krippendorff α")
    test_krippendorff_alpha_known_values()
    print("Test muestreo")
    test_sampling()
    test_small_corpus_exhaustive()
    print("Test end-to-end")
    test_end_to_end()
    test_disagreement_case()
    print("Test acuerdo intra-experto en reporte")
    test_inter_annotator_agreement_via_compare()
    print("Test refusal rate")
    test_refusal_propagates_to_metrics_and_report()
    print("\nTodos los tests OK.")
