"""DEMO — Rellena las hojas de anotación con valores SINTÉTICOS.

ADVERTENCIA: este script NO produce validación humana real. Genera
anotaciones simuladas (cada "anotador" coincide con el juez con cierta
probabilidad, con ruido y desacuerdo entre anotadores) para:
  (a) verificar end-to-end que compute_agreement.py produce kappa de Cohen,
      Fleiss y Krippendorff, y
  (b) tener una PLANTILLA rellena de cómo quedarían las tablas.

Los valores resultantes son de DEMOSTRACIÓN y deben sustituirse por la
anotación real de los tres integrantes antes de entregar la memoria.

Uso:
    python scripts/fill_demo_annotations.py <carpeta_run>
    (por defecto: la más reciente con subcarpeta review/)
"""

from __future__ import annotations

import json
import random
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent

# Perfiles de "fiabilidad" por anotador: (p_acuerdo_unsafe, p_acuerdo_fallo).
PROFILES = {
    "diego": (0.93, 0.88),
    "gontzal": (0.85, 0.80),
    "jose": (0.90, 0.85),
}
FAILS = ["blind_obedience", "normative_collapse", "internal_contradiction"]


def main() -> None:
    if len(sys.argv) > 1:
        run_dir = Path(sys.argv[1])
        if not run_dir.is_absolute():
            run_dir = ROOT / run_dir
    else:
        runs = sorted((ROOT / "results").glob("*/review"), key=lambda p: p.stat().st_mtime)
        if not runs:
            raise SystemExit("No hay ninguna carpeta results/*/review/. Ejecuta antes export_for_review.py.")
        run_dir = runs[-1].parent

    results_jsonl = run_dir / "results.jsonl"
    review = run_dir / "review"
    verdicts = {}
    for line in results_jsonl.open(encoding="utf-8"):
        r = json.loads(line)
        verdicts[r["run_id"]] = r["verdict"]
    print(f"Run: {run_dir.name} | {len(verdicts)} veredictos del juez")

    random.seed(7)
    for aid, (p_unsafe, p_fail) in PROFILES.items():
        path = review / f"annotations_{aid}.xlsx"
        wb = load_workbook(path)
        ws = wb["annotations"]
        n = 0
        for row in ws.iter_rows(min_row=3):
            run_id = row[0].value
            if not run_id:
                continue
            v = verdicts.get(str(run_id))
            if not v:
                continue
            ju = bool(v["is_unsafe_helpful"])
            jf = v.get("dominant_failure") or "none"
            jr = bool(v.get("is_refusal", False))

            # is_unsafe_helpful: coincide con el juez con prob p_unsafe.
            au = ju if random.random() < p_unsafe else (not ju)
            # dominant_failure coherente con au.
            if not au:
                af = "none"
            elif random.random() < p_fail and jf != "none":
                af = jf
            else:
                af = random.choice(FAILS)
            # is_refusal: por construcción, si es unsafe no es negativa pura.
            ar = False if au else (jr if random.random() < 0.85 else (not jr))

            rn = row[0].row
            ws.cell(row=rn, column=8, value=str(au).upper())   # H is_unsafe_helpful
            ws.cell(row=rn, column=9, value=af)                # I dominant_failure
            ws.cell(row=rn, column=11, value=str(ar).upper())  # K is_refusal
            n += 1
        wb.save(path)
        print(f"  rellenado (DEMO) {path.name}: {n} filas")

    print("\n[DEMO] Valores sintéticos. Sustituir por anotación real antes de entregar.")


if __name__ == "__main__":
    main()
